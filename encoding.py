import json
import openpyxl

# Load json files
with open('_high.json') as f:
    high_list = json.load(f)

with open('bu.json') as f:
    bu_list = json.load(f)

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Write headers
ws.cell(row=1, column=1).value = "Business Unit"
ws.cell(row=1, column=2).value = "Username"
ws.cell(row=1, column=3).value = "Email"

# Loop through both lists and write to Excel
for i in range(len(high_list)):
    for j in range(len(bu_list)):
        ws.append([bu_list[j], high_list[i][0], high_list[i][1]])

# Save workbook
wb.save("output.xlsx")
